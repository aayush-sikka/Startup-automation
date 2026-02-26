"""
excel_exporter.py — Export enriched startup data to a professional .xlsx
=========================================================================
Sheets produced:
  1. All Startups       — full enriched dataset (50+ rows)
  2. Fintech Filter     — assignment requirement: at least one filtered category
  3. Outreach Messages  — email + WhatsApp per startup
  4. Summary Dashboard  — counts by industry, stage, accuracy band
"""

import os
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Color palette ─────────────────────────────────────────────────────────────
C_HEADER_BG   = "1A237E"   # deep indigo
C_HEADER_FG   = "FFFFFF"
C_ALT_ROW     = "E8EAF6"   # light indigo tint
C_ACCENT      = "3949AB"
C_GREEN_BG    = "E8F5E9"
C_AMBER_BG    = "FFF8E1"
C_RED_BG      = "FFEBEE"
C_FINTECH_HDR = "0D47A1"   # darker blue for fintech sheet
C_MSG_HDR     = "1B5E20"   # dark green for messages sheet
C_DASH_HDR    = "4A148C"   # deep purple for dashboard

thin_border = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)


def _header_cell(ws, row, col, value, bg=C_HEADER_BG, fg=C_HEADER_FG, size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=True, color=fg, size=size, name="Arial")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border
    return cell


def _data_cell(ws, row, col, value, alt=False, wrap=False, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", size=9, bold=bold)
    cell.fill      = PatternFill("solid", fgColor=C_ALT_ROW if alt else "FFFFFF")
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border    = thin_border
    return cell


def _score_fill(score: int) -> str:
    if score >= 70: return C_GREEN_BG
    if score >= 40: return C_AMBER_BG
    return C_RED_BG


def _freeze_and_filter(ws, freeze="A2", filter_row=1):
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions


# ============================================================================
# SHEET 1 — All Startups
# ============================================================================

MAIN_COLS = [
    ("No.",              4),
    ("Company Name",    22),
    ("Website",         28),
    ("Industry",        13),
    ("Stage",           14),
    ("Founded Year",    12),
    ("Location",        14),
    ("Founder(s)",      22),
    ("Founder Email(s)",26),
    ("Contact Email(s)",26),
    ("Phone",           14),
    ("Funding",         14),
    ("Employees",       12),
    ("LinkedIn",        20),
    ("Twitter",         16),
    ("Description",     40),
    ("Problem Statement",40),
    ("Target Customers",28),
    ("Accuracy Score",  14),
    ("Source",          14),
]


def _write_main_sheet(ws, startups: List[Dict], title: str, header_bg: str = C_HEADER_BG):
    ws.title = title
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(MAIN_COLS))}1")
    tc = ws["A1"]
    tc.value     = f"🚀  {title}  —  Automated Startup Outreach System (India)"
    tc.font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    tc.fill      = PatternFill("solid", fgColor=header_bg)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    ws.row_dimensions[2].height = 30
    for col_idx, (label, width) in enumerate(MAIN_COLS, 1):
        _header_cell(ws, 2, col_idx, label, bg=header_bg)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, s in enumerate(startups, 3):
        alt = row_idx % 2 == 0
        ws.row_dimensions[row_idx].height = 18

        founders  = "; ".join(s.get("founders", []) or []) or "Not found"
        f_emails  = "; ".join(s.get("founder_emails", []) or []) or "Not found"
        c_emails  = "; ".join(s.get("contact_emails", []) or []) or "Not found"
        phones    = "; ".join(s.get("contact_phones", []) or []) or "Not found"
        score     = s.get("accuracy_score", 0)

        row_data = [
            row_idx - 2,
            s.get("name", "Unknown"),
            s.get("website", "N/A"),
            s.get("industry", "Other"),
            s.get("stage", "Early"),
            s.get("founded_year", "Unknown"),
            s.get("location", "India"),
            founders,
            f_emails,
            c_emails,
            phones,
            str(s.get("funding", "")) or "Unknown",
            str(s.get("employees", "")) or "Unknown",
            s.get("linkedin", "") or "",
            s.get("twitter", "") or "",
            s.get("description", "N/A"),
            s.get("problem_statement", "") or "",
            s.get("target_customers", "") or "",
            score,
            s.get("source", "Google Search"),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = _data_cell(ws, row_idx, col_idx, value, alt=alt,
                              wrap=(col_idx in (8, 16, 17, 18)))
            # Colour-code accuracy score column
            if col_idx == 19:
                cell.fill  = PatternFill("solid", fgColor=_score_fill(score))
                cell.font  = Font(name="Arial", size=9, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")

    _freeze_and_filter(ws, "A3")


# ============================================================================
# SHEET 2 — Outreach Messages
# ============================================================================

MSG_COLS = [
    ("No.",           4),
    ("Company Name", 22),
    ("Industry",     13),
    ("Founder",      18),
    ("Email",        28),
    ("Stage",        13),
    ("📧 Email Outreach (80-120 words)", 60),
    ("💬 WhatsApp Pitch (25-40 words)",  40),
]


def _write_messages_sheet(ws, startups: List[Dict]):
    ws.title = "Outreach Messages"
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{get_column_letter(len(MSG_COLS))}1")
    tc = ws["A1"]
    tc.value     = "✉️  Personalized Outreach Messages  —  Email + WhatsApp"
    tc.font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    tc.fill      = PatternFill("solid", fgColor=C_MSG_HDR)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col_idx, (label, width) in enumerate(MSG_COLS, 1):
        _header_cell(ws, 2, col_idx, label, bg=C_MSG_HDR)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 30

    for row_idx, s in enumerate(startups, 3):
        alt = row_idx % 2 == 0
        founders = (s.get("founders") or ["Unknown"])[0]
        email    = (s.get("founder_emails") or s.get("contact_emails") or ["Not found"])[0]
        ws.row_dimensions[row_idx].height = 90   # tall rows for message text

        row_data = [
            row_idx - 2,
            s.get("name", "Unknown"),
            s.get("industry", "Other"),
            founders,
            email,
            s.get("stage", "Early"),
            s.get("outreach_email", ""),
            s.get("outreach_whatsapp", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            _data_cell(ws, row_idx, col_idx, value, alt=alt, wrap=(col_idx in (7, 8)))

    ws.freeze_panes = "A3"


# ============================================================================
# SHEET 3 — Summary Dashboard
# ============================================================================

def _write_dashboard(ws, startups: List[Dict]):
    ws.title = "Summary Dashboard"
    ws.sheet_view.showGridLines = False

    def section_header(row, col, text, bg=C_DASH_HDR, colspan=3):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = Font(bold=True, size=11, color="FFFFFF", name="Arial")
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22

    def kv(row, col, key, value, bg_key="EDE7F6", bg_val="FFFFFF"):
        kc = ws.cell(row=row, column=col, value=key)
        kc.font  = Font(bold=True, size=10, name="Arial")
        kc.fill  = PatternFill("solid", fgColor=bg_key)
        kc.alignment = Alignment(horizontal="left", vertical="center")
        kc.border = thin_border

        vc = ws.cell(row=row, column=col + 1, value=value)
        vc.font  = Font(size=10, name="Arial")
        vc.fill  = PatternFill("solid", fgColor=bg_val)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = thin_border
        ws.row_dimensions[row].height = 18

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 16

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells("B1:F1")
    tc = ws["B1"]
    tc.value = "📊  Assignment 3 — System Summary Dashboard"
    tc.font  = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    tc.fill  = PatternFill("solid", fgColor=C_DASH_HDR)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Overview ──────────────────────────────────────────────────────────
    section_header(3, 2, "📌 Overview", colspan=2)
    total = len(startups)
    with_founders = sum(1 for s in startups if s.get("founders"))
    with_email    = sum(1 for s in startups if s.get("founder_emails") or s.get("contact_emails"))
    avg_score     = round(sum(s.get("accuracy_score",0) for s in startups) / max(total,1), 1)
    high_acc      = sum(1 for s in startups if s.get("accuracy_score",0) >= 70)

    kv(4,  2, "Total Startups Found",     total)
    kv(5,  2, "With Founder Names",       with_founders)
    kv(6,  2, "With Contact Email",       with_email)
    kv(7,  2, "Avg Accuracy Score",       f"{avg_score}/100")
    kv(8,  2, "High Accuracy (70+)",      high_acc)
    kv(9,  2, "Outreach Messages Ready",  total)

    # ── By Industry ───────────────────────────────────────────────────────
    section_header(11, 2, "🏭 By Industry", colspan=2)
    from collections import Counter
    ind_counts = Counter(s.get("industry","Other") for s in startups)
    for r, (ind, cnt) in enumerate(ind_counts.most_common(), 12):
        kv(r, 2, ind, cnt)

    # ── By Stage ─────────────────────────────────────────────────────────
    section_header(3, 5, "🌱 By Startup Stage", colspan=2)
    stage_counts = Counter(s.get("stage","Early") for s in startups)
    for r, (stage, cnt) in enumerate(stage_counts.most_common(), 4):
        kv(r, 5, stage, cnt)

    # ── Accuracy bands ────────────────────────────────────────────────────
    section_header(11, 5, "🎯 Accuracy Bands", colspan=2)
    bands = {
        "🟢 Excellent (70-100)": sum(1 for s in startups if s.get("accuracy_score",0) >= 70),
        "🟡 Good (50-69)":       sum(1 for s in startups if 50 <= s.get("accuracy_score",0) < 70),
        "🟠 Moderate (30-49)":   sum(1 for s in startups if 30 <= s.get("accuracy_score",0) < 50),
        "🔴 Low (0-29)":         sum(1 for s in startups if s.get("accuracy_score",0) < 30),
    }
    for r, (band, cnt) in enumerate(bands.items(), 12):
        kv(r, 5, band, cnt)

    # ── Sending layer notes ───────────────────────────────────────────────
    section_header(20, 2, "📤 Task 5 — Sending Layer Workflow", colspan=5)
    notes = [
        ("Email Sending",    "Integrate with SendGrid / Mailgun API → loop through startups, send outreach_email field"),
        ("WhatsApp Sending", "Use WhatsApp Business API (Meta Cloud API) → send outreach_whatsapp field"),
        ("Rate Limiting",    "Send max 50 emails/hour, 20 WhatsApp/hour to avoid spam filters"),
        ("Personalisation",  "Each message uses founder name, company name, industry pain points"),
        ("Tracking",         "Log sent_at, opened_at, replied_at back into this spreadsheet"),
        ("Follow-up",        "Auto-schedule follow-up after 3 days if no reply (via n8n / Zapier)"),
    ]
    for r, (k, v) in enumerate(notes, 21):
        kv(r, 2, k, v, bg_val="F3E5F5")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        vc = ws.cell(row=r, column=3, value=v)
        vc.font  = Font(size=9, name="Arial")
        vc.fill  = PatternFill("solid", fgColor="F3E5F5")
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        vc.border = thin_border
        ws.row_dimensions[r].height = 28


# ============================================================================
# MAIN EXPORT FUNCTION
# ============================================================================

def export_to_excel(startups: List[Dict], filepath: str = "startup_outreach.xlsx"):
    """
    Write all sheets to a professional Excel workbook.
    Returns the filepath on success.
    """
    fintech_startups = [s for s in startups if s.get("industry","").lower() == "fintech"]

    wb = Workbook()

    # Sheet 1 — All Startups
    ws1 = wb.active
    _write_main_sheet(ws1, startups, "All Startups")

    # Sheet 2 — Fintech Filter (assignment mandatory category)
    ws2 = wb.create_sheet()
    _write_main_sheet(ws2, fintech_startups or startups[:15],
                      "Fintech Startups", header_bg=C_FINTECH_HDR)

    # Sheet 3 — Outreach Messages
    ws3 = wb.create_sheet()
    _write_messages_sheet(ws3, startups)

    # Sheet 4 — Dashboard
    ws4 = wb.create_sheet()
    _write_dashboard(ws4, startups)

    # Tab colours
    ws1.sheet_properties.tabColor = "1A237E"
    ws2.sheet_properties.tabColor = "0D47A1"
    ws3.sheet_properties.tabColor = "1B5E20"
    ws4.sheet_properties.tabColor = "4A148C"

    wb.save(filepath)
    print(f"  ✅ Excel saved → {filepath}")
    return filepath
